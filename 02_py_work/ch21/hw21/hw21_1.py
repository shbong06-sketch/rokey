# hw21_1.py

# 5.
import json

py_data = {
    "name": "홍길동",
    "age": 25,
    "city": "서울"
    }

# with open('data.json', 'w') as file:
#     json.dump(py_data, file, indent=4)

# 6.
import os
print(f"현재 작업 디렉로리: {os.getcwd()}")
os.chdir(r".\ch21\manage_file")
print(f"변경 작업 디렉로리: {os.getcwd()}")

print(f"디렉토리 목록: {os.listdir(".")}")

print("-------------------------")
# 7. 
from pathlib import Path

def list_text_file(folder_path):
    f_list = list()
    folder = Path(folder_path)
    for file in folder.iterdir():
        if file.is_file() and file.suffix == '.txt':
            f_list.append(file)
    return f_list

folder = r"C:\rokey\py_work\ch21\manage_file\folder1"
print("1.txt 파일 목록:", list_text_file(folder))

print("-----------------------")
# 8.
from pathlib import Path
new_folder = Path("new_folder")
new_folder.mkdir(exist_ok=True)

print("-----------------------")
# 9.

# from openpyxl import Workbook

# wb = Workbook()
# sheet = wb.active
# wb.save("data.xlsx")

print("-----------------------")
# 10.
from openpyxl import load_workbook

wb = load_workbook("data.xlsx")

sheet = wb.active

sheet['A1'] = 'Hello'
wb.save('data.xlsx')
